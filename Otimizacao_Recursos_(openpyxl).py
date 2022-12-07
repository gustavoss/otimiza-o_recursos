from openpyxl import load_workbook
import sys

sys.stdout = open("output.txt", "w")

#carrega planilha
wb = load_workbook(filename = 'SJP.xlsm', data_only=True)

#conta número de abas
sheets_number = len(wb.worksheets)
print("Quantidade de abas: " + str(sheets_number))

last_row = []
last_col = []

#encontra ultima linha e coluna preenchida de todas as abas
for i in range(0, sheets_number):
    last_row.append(wb.worksheets[i].max_row)
    last_col.append(wb.worksheets[i].max_column)

#cria matriz de volumes
Volumes = [[0 for i in range(1, last_col[2] + 1)] for j in range(1, last_row[2])]

#atribuí valores à matriz de volumes
for i in range(0, last_row[2] - 1):
    for j in range(0, last_col[2]):
        Volumes[int(i)][int(j)] = wb.worksheets[2].cell(row=int(i + 1), column=int(j + 1)).value

#cria matriz de ocupações
Ocupacoes = [[0 for i in range(1, last_col[2] + 1)] for j in range(1, last_row[2])]

#atribuí valores à matriz de ocupações
for i in range(0, last_row[3]):
    for j in range(0, last_col[3]):
        Ocupacoes[int(i)][int(j)] = wb.worksheets[3].cell(row=int(i + 1), column=int(j + 1)).value

for i in range(2, last_row[0]):
    for j in range(2, last_row[5]):

        #procura SKU, novo recurso e se o novo recurso é mesma tecnologia
        if wb.worksheets[0].cell(row=i, column=3).value == wb.worksheets[5].cell(row=j, column=2).value \
        and wb.worksheets[0].cell(row=i, column=2).value != wb.worksheets[5].cell(row=j, column=3).value \
        and wb.worksheets[0].cell(row=i, column=11).value == wb.worksheets[5].cell(row=j, column=7).value:
            #print("Outro recurso encontrado!")

            ocupacao_atual = 0
            ocupacao_nova = 0

            #check capacidade do recurso atual
            for x in range(0, last_row[3]):
                for y in range(0, last_col[3]):
                    if wb.worksheets[0].cell(row=i, column=2).value == Ocupacoes[x][1] and wb.worksheets[0].cell(row=i, column=9).value == Ocupacoes[0][y]:
                        ocupacao_atual = Ocupacoes[x][y]

            # check capacidade do recurso novo
            for x in range(0, last_row[3]):
                for y in range(0, last_col[3]):
                    if wb.worksheets[5].cell(row=j, column=3).value == Ocupacoes[x][1] and wb.worksheets[0].cell(row=i, column=9).value == Ocupacoes[0][y]:
                        ocupacao_nova = Ocupacoes[x][y]

            #check se recurso novo tem melhor capacidade que o atual
            if float(ocupacao_nova) < float(ocupacao_atual) and float(ocupacao_atual) > 0.7:
                #print("Recurso melhor encontrado para ordem " + str(wb.worksheets[0].cell(row=i, column=5).value) + "!")

                recurso_antigo = wb.worksheets[0].cell(row=i, column=2).value
                recurso_novo = wb.worksheets[5].cell(row=j, column=3).value

                flag_ausencia_rec_antigo = True
                flag_ausencia_rec_novo = True

                for x in range(0, last_row[3]):
                    if recurso_antigo == Volumes[x][1]:
                        flag_ausencia_rec_antigo = False
                        continue

                for x in range(0, last_row[3]):
                    if recurso_novo == Volumes[x][1]:
                        flag_ausencia_rec_novo = False
                        continue

                if flag_ausencia_rec_antigo:
                    break

                if flag_ausencia_rec_novo:
                    break

                if i == 1227:
                    teste = "teste"
                #atualizar volume do recurso antigo
                flag_recurso_antigo = False
                flag_volume = False
                for x in range(1, last_row[3]):
                    if flag_recurso_antigo:
                        break
                    for y in range(3, last_col[3] - 1):
                        #print(recurso_antigo)
                        #print(Ocupacoes[x][1])
                        #print(wb.worksheets[0].cell(row=i, column=9).value)
                        #print(Ocupacoes[0][y])
                        #print(Volumes[x][y])
                        #print(wb.worksheets[0].cell(row=i, column=7).value)
                        if recurso_antigo == Ocupacoes[x][1] and wb.worksheets[0].cell(row=i, column=9).value == Ocupacoes[0][y] \
                        and (Volumes[x][y] - wb.worksheets[0].cell(row=i, column=7).value) < 0:
                            flag_volume = True
                        if recurso_antigo == Ocupacoes[x][1] and wb.worksheets[0].cell(row=i, column=9).value == Ocupacoes[0][y] \
                        and (Volumes[x][y] - wb.worksheets[0].cell(row=i, column=7).value) >= 0:
                            Volumes[x][y] = Volumes[x][y] - wb.worksheets[0].cell(row=i, column=7).value
                            wb.worksheets[2].cell(row=x + 1, column= y + 1).value = Volumes[x][y]
                            #nova ocupacao do recurso antigo
                            nova_ocupacao_rec_antigo = Volumes[x][y] / wb.worksheets[4].cell(row=x+1, column=y + 1).value
                            wb.worksheets[3].cell(row=x + 1, column=y + 1).value = (nova_ocupacao_rec_antigo)
                            print(str(recurso_antigo) + ": " + "-" + str(wb.worksheets[0].cell(row=i, column=7).value))
                            flag_recurso_antigo = True
                            if flag_recurso_antigo:
                                break
                if flag_volume:
                    break

                # atualizar volume do recurso novo
                flag_recurso_novo = False

                for x in range(1, last_row[3]):
                    if flag_recurso_novo:
                        break
                    for y in range(3, last_col[3] - 1):
                        #print(recurso_novo)
                        #print(Ocupacoes[x][1])
                        #print(wb.worksheets[0].cell(row=i, column=9).value)
                        #print(Ocupacoes[0][y])
                        if recurso_novo == Ocupacoes[x][1] and wb.worksheets[0].cell(row=i, column=9).value == Ocupacoes[0][y]:
                            Volumes[x][y] = Volumes[x][y] + wb.worksheets[0].cell(row=i, column=7).value
                            #wb.worksheets[2].cell(row=x + 1, column=y + 1).value = Volumes[x][y]
                            # nova ocupacao do recurso antigo
                            nova_ocupacao_rec_novo = Volumes[x][y] / wb.worksheets[4].cell(row=x + 1, column=y + 1).value
                            wb.worksheets[3].cell(row=x + 1, column=y + 1).value = (nova_ocupacao_rec_novo)
                            print(str(recurso_novo) + ": " + "+" + str(wb.worksheets[0].cell(row=i, column=7).value))
                            flag_recurso_novo = True
                            if flag_recurso_novo:
                                break

                wb.worksheets[0].cell(row=i, column=14).value = wb.worksheets[0].cell(row=i, column=2).value
                wb.worksheets[0].cell(row=i, column=13).value = "Equalizado"
                wb.worksheets[0].cell(row=i, column=2).value = str(wb.worksheets[5].cell(row=j, column=3).value)
                print("Linha " + str(i) + " Equalizada!")
                break

# salva planilha
wb.save('SJP.xlsx')
sys.stdout.close()


